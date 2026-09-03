"""User-entered source records and a cumulative, locally stored Excel ledger."""
from pathlib import Path
from datetime import datetime
import portalocker
import hashlib
import json
import os
import tempfile
import uuid
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from notices import SHORT_NOTICE,disclaimer_text

FIELDS = [
    ("evidence_id", "证据编号"), ("collected", "是否取证"), ("platform", "平台"),
    ("account_type", "账号类别"), ("author", "发布者账号或作者"), ("uid", "账号 ID / UID"),
    ("published_at", "发布时间（原页面显示）"), ("location", "发布地点 / IP 属地（原页面显示）"),
    ("title", "视频标题"), ("description", "视频文案"), ("url", "视频链接"),
    ("issue", "涉及内容 / 疑似侵权点（填写人描述）"), ("capture_at", "取证日期及时间（填写）"),
    ("capture_location", "取证地点（填写）"), ("evidence_filename", "取证文件名（填写）"),
    ("rights_guard_used", "是否使用权利卫士取证"),
    ("rights_guard_record_name", "权利卫士内记录名称"),
    ("rights_guard_evidence_name", "权利卫士证据名称"),
    ("rights_guard_evidence_id", "权利卫士证据编号 / 存证编号"),
    ("rights_guard_capture_at", "权利卫士取证 / 固化时间"),
    ("rights_guard_original_filename", "权利卫士原始取证文件名"),
    ("rights_guard_certificate_name", "权利卫士认证证书文件名"),
    ("rights_guard_verification_url", "权利卫士证据分享 / 验证链接"),
    ("trusted_timestamp_used", "是否取得第三方可信时间戳"),
    ("timestamp_service", "时间戳 / 存证服务名称"),
    ("timestamp_certificate_no", "时间戳证书编号 / 序列号"),
    ("timestamp_issued_at", "可信时间戳签发时间"),
    ("timestamp_certificate_file", "可信时间戳证书文件名"),
    ("timestamp_bound_hash", "证书对应文件哈希 / 数字指纹"),
    ("timestamp_verification_url", "可信时间戳验证链接"),
    ("views", "浏览量"), ("likes", "点赞数"), ("favorites", "收藏数"),
    ("shares", "转发数"), ("comments", "评论数"), ("metrics_at", "传播数据观察时间（填写）"),
    ("platform_status", "平台处理状态"), ("platform_details", "平台处理情况 / 投诉编号"),
    ("platform_handled_at", "平台处理时间（填写）"), ("notes", "备注"),
]
AUTOMATIC = [
    ("record_id", "记录 ID"), ("export_started_at", "导出开始时间（本机含时区）"),
    ("export_finished_at", "导出完成时间（本机含时区）"), ("original_file", "原视频文件"),
    ("sha256", "原视频 SHA-256"), ("range", "处理范围（原视频时间）"),
    ("output_folder", "输出文件夹"), ("generated_files", "生成文件名"),
    ("attachment_names", "评论区等附件文件名"), ("screenshot_count", "截图数量"),
]
FIELD_NOTES = [
    SHORT_NOTICE,
    "取证字段属于使用者填写的整理标签，不构成认证。使用者应核实内容权利和来源并保护隐私。",
    "所有来源、账号、发布信息、传播数据和平台状态均由使用者填写，程序不从网页自动抓取。",
    "取证日期和时间由使用者填写；导出时间是本机系统时间（含时区），两者不是同一个事件，也不是可信时间戳。",
    "“权利卫士”和可信时间戳字段仅记录使用者从第三方服务取得并核对的信息；程序不连接、验证或代替第三方服务。",
    "权利卫士记录名称、证据名称、编号、固化时间、证书文件名和验证链接应按第三方页面或证书原样填写。",
    "可信时间戳签发时间、证书编号、文件哈希和验证链接应按证书原样填写；不要用本机当前时间代替。",
    "未知、不可见或未核实的传播数据留空，可原样填写 1.2万 / 不可见。空白不等于零。",
    "是否取证不会因生成文件自动改为是，请自行核对。疑似侵权点是填写人的描述。",
    "原视频哈希用于核对文件一致性。保留原视频；输出截图和文档属于加工文件。",
    "发布地点与取证地点分别记录，不从视频或电脑位置推断。",
    "台账每次只在末尾追加新行，不按证据编号替换旧行；旧版缺少的字段自动加到最右侧，原有行、手工修改和自定义列保留。",
    "可在 Excel 中继续编辑已有记录。程序追加前，请先保存并关闭 Excel 中的台账。",
    "长字段设置了换行，可在 Excel 中调整行高查看全文。",
]


def local_time():
    return datetime.now().astimezone().isoformat(timespec="seconds")


def new_evidence_id():
    return "EV-" + datetime.now().strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:4].upper()


def fingerprint(path, check=lambda: None):
    digest = hashlib.sha256()
    with Path(path).open("rb") as f:
        while chunk := f.read(4*1024*1024):
            check()
            digest.update(chunk)
    return digest.hexdigest()


def make_record(fields, **automatic):
    result = {key: str(fields.get(key, "")).strip() for key, _ in FIELDS}
    result["evidence_id"] = result["evidence_id"] or new_evidence_id()
    result["collected"] = result["collected"] or "待核对"
    result.update(automatic)
    result["record_id"] = automatic.get("record_id") or uuid.uuid4().hex
    return result


def set_text(cell, value):
    # Force user input to literal strings (UIDs, leading zeroes, and '=...' text).
    # Unknown counts remain blank; no accidental formulas or inferred zeroes.
    text = "" if value is None else str(value)
    if len(text) > 32767:
        raise ValueError("单个表格字段超过 Excel 的 32767 字符限制，请缩短文案或备注。")
    cell.value = text
    cell.data_type = "s"
    cell.number_format = "@"


def style_header(sheet, column, title):
    cell = sheet.cell(1, column, title)
    cell.fill = PatternFill("solid", fgColor="147C67")
    cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.column_dimensions[get_column_letter(column)].width = max(24, min(42, len(title) * 2 + 4))


def ensure_schema(sheet, schema):
    """Add new columns on the right without moving or changing existing cells."""
    titles = [title for _, title in schema]
    existing = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    if not any(value is not None for value in existing):
        for column, title in enumerate(titles, 1):
            style_header(sheet, column, title)
        sheet.row_dimensions[1].height = 34
        sheet.freeze_panes = "D2"
        return

    known = [value for value in existing if value in titles]
    if not known:
        raise ValueError("所选文件不是可识别的 VideoNotes 取证台账；为保护原内容，程序没有保存。")
    duplicates = {title for title in titles if existing.count(title) > 1}
    if duplicates:
        raise ValueError("取证台账存在重复表头：" + "、".join(sorted(duplicates)) + "。为保护原内容，程序没有保存。")

    # A schema upgrade is append-only: old columns, custom columns, old rows,
    # formulas, styles and manual edits keep their original positions.
    for title in titles:
        if title not in existing:
            column = sheet.max_column + 1
            style_header(sheet, column, title)
            existing.append(title)


def add_record(workbook, record):
    sheet = workbook["取证台账"] if "取证台账" in workbook.sheetnames else workbook.create_sheet("取证台账")
    schema = FIELDS + AUTOMATIC
    ensure_schema(sheet, schema)
    headers = {sheet.cell(1,c).value: c for c in range(1,sheet.max_column+1)}
    row = sheet.max_row + 1
    for key, title in schema:
        cell = sheet.cell(row,headers[title])
        set_text(cell,record.get(key,""))
        cell.font = Font(name="Microsoft YaHei", size=11)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="EFF7F3")
    sheet.row_dimensions[row].height = 75
    sheet.auto_filter.ref = sheet.dimensions
    if "免责声明" not in workbook.sheetnames:
        notice=workbook.create_sheet("免责声明")
        notice.column_dimensions["A"].width=105
        for text in disclaimer_text().splitlines():
            if not text.strip(): continue
            index=notice.max_row+1 if notice.cell(1,1).value is not None else 1
            cell=notice.cell(index,1)
            set_text(cell,text)
            cell.alignment=Alignment(wrap_text=True,vertical="top")
            cell.font=Font(name="Microsoft YaHei",size=11,bold=text.startswith("#"))
            notice.row_dimensions[index].height=max(25,((len(text)+65)//66)*17)
    info = workbook["字段说明"] if "字段说明" in workbook.sheetnames else workbook.create_sheet("字段说明")
    existing_notes = {str(info.cell(row, 1).value) for row in range(1, info.max_row + 1) if info.cell(row, 1).value}
    for note in FIELD_NOTES:
        if note not in existing_notes:
            info.append([note])
    info.column_dimensions["A"].width = 110
    for cells in info:
        cells[0].alignment = Alignment(wrap_text=True,vertical="top")
        info.row_dimensions[cells[0].row].height = 42


def workbook_for(record):
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_record(workbook,record)
    return workbook


def write_record(record, folder):
    folder = Path(folder)
    (folder/"取证信息.json").write_text(json.dumps(record,ensure_ascii=False,indent=2),encoding="utf-8")
    workbook_for(record).save(folder/"取证信息.xlsx")


def append_ledger(record, output):
    output = Path(output)
    output.mkdir(parents=True,exist_ok=True)
    path = output/"取证台账.xlsx"
    with portalocker.Lock(str(output/".取证台账.lock"),mode="a",timeout=15):
        if (output/"~$取证台账.xlsx").exists():
            raise ValueError("请先保存并关闭 Excel 中打开的取证台账，再追加记录。")
        before = path.stat().st_mtime_ns if path.exists() else None
        book = load_workbook(path) if path.exists() else Workbook()
        if not path.exists():
            book.remove(book.active)
        add_record(book,record)
        fd, temp = tempfile.mkstemp(suffix=".xlsx",prefix=".台账-",dir=output)
        os.close(fd)
        try:
            book.save(temp)
            if (path.stat().st_mtime_ns if path.exists() else None) != before:
                raise ValueError("台账在保存期间被其他程序修改，请关闭 Excel 后重试。")
            os.replace(temp,path)
        finally:
            Path(temp).unlink(missing_ok=True)
            book.close()
    return path
