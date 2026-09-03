import sys
import unittest
import tempfile
import threading
import json
from pathlib import Path
sys.path.insert(0,str(Path(__file__).resolve().parents[1]))
from PIL import Image,ImageDraw
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
from engine import ROOT,Options,Segment,process,run,executable,write_srt
from evidence import make_record,append_ledger,fingerprint
from visual import Detector
import numpy as np


class Upgrade(unittest.TestCase):
    def test_gesture_overlay_scene_and_spreadsheet(self):
        with tempfile.TemporaryDirectory() as td:
            root=Path(td)
            for i in range(8):
                im=Image.new("RGB",(640,360),"#c5d7d0")
                draw=ImageDraw.Draw(im)
                draw.rectangle((70,60,170,200),fill="#383333")
                if i in (1,2):
                    draw.rectangle((170,80,260+i*25,115),fill="#383333")
                if i==3: im=Image.new("RGB",(640,360),"#913994")
                if i==5: draw.rectangle((390,65,455,130),fill="#ff2400")
                im.save(root/f"{i}.png")
            (root/"frames.txt").write_text("".join(f"file '{i}.png'\nduration 0.5\n" for i in range(8))+"file '7.png'\n")
            video=root/"motion.mp4"
            run([executable("ffmpeg"),"-v","error","-y","-f","concat","-safe","0","-i",root/"frames.txt","-t","4","-r","25","-c:v","libx264","-pix_fmt","yuv420p",video])
            write_srt([Segment(0,4,"字幕始终不变")],root/"same.srt")
            fields={"evidence_id":"EV-DEMO-001","uid":"00001234567890123456789","description":"=1+1","views":"","likes":"1.2万","platform":"B站","collected":"否","capture_at":"2026-09-03 10:15 +08:00","title":"人物动作与插入表情包测试"}
            output,report=process(Options(str(video),str(root/"out"),subtitle=str(root/"same.srt"),word_source="subtitles",interval=.2,
                evidence=fields,attachments=[str(root/"5.png")]))
            times=[r["capture"] for r in report["segments"] if "画面" in r["reason"]]
            print("Visual changes:",times)
            self.assertTrue(any(.4<=t<=.8 for t in times),"gesture missed")
            self.assertTrue(any(1.4<=t<=1.8 for t in times),"cut missed")
            self.assertTrue(any(2.4<=t<=2.8 for t in times),"small overlay missed")
            self.assertGreaterEqual(len(times),6)
            self.assertEqual(report["transcript_segments"],1)
            record=report["evidence"]
            self.assertEqual(record["sha256"],fingerprint(video))
            self.assertEqual(record["collected"],"否")
            self.assertEqual(record["capture_at"],fields["capture_at"])
            self.assertTrue((output/"评论区等附件/001_5.png").is_file())
            for file in (output/"取证信息.xlsx",root/"out/取证台账.xlsx"):
                book=load_workbook(file)
                sheet=book["取证台账"]
                head={sheet.cell(1,c).value:c for c in range(1,sheet.max_column+1)}
                self.assertEqual(sheet.cell(2,head["账号 ID / UID"]).value,fields["uid"])
                self.assertEqual(sheet.cell(2,head["视频文案"]).data_type,"s")
                self.assertIsNone(sheet.cell(2,head["浏览量"]).value)
                book.close()
            # Silent, text-free footage still produces a visual PDF.
            silent,report2=process(Options(str(video),str(root/"out"),mode="speech",interval=.2))
            self.assertEqual(report2["word_source"],"无可用语音或字幕")
            self.assertGreater(report2["screenshots"],1)
            book=load_workbook(root/"out/取证台账.xlsx")
            self.assertEqual(book["取证台账"].max_row,3)
            book.close()
            import shutil
            sample=ROOT/"tests/v2-demo"
            sample.mkdir(exist_ok=True)
            for name in ("字幕截图.pdf","取证信息.xlsx","取证信息.json","处理记录.json"):
                shutil.copy2(output/name,sample/name)
            shutil.copy2(root/"out/取证台账.xlsx",sample/"取证台账.xlsx")

    def test_noise_vs_small_change(self):
        frame=np.zeros((180,320,3),np.uint8)+100
        detector=Detector()
        self.assertTrue(detector.accept(0,frame))
        self.assertFalse(detector.accept(.5,frame+1))
        changed=frame.copy()
        changed[30:55,50:80]=255
        self.assertTrue(detector.accept(1,changed))

    def test_ledger_preserves_manual_edit(self):
        with tempfile.TemporaryDirectory() as td:
            path=append_ledger(make_record({"title":"第一条"}),td)
            book=load_workbook(path)
            book["取证台账"]["A2"]="手工修改的编号"
            book.save(path)
            book.close()
            append_ledger(make_record({"title":"第二条"}),td)
            book=load_workbook(path)
            self.assertEqual(book["取证台账"]["A2"].value,"手工修改的编号")
            self.assertEqual(book["取证台账"].max_row,3)
            book.close()

    def test_old_ledger_is_extended_without_moving_or_overwriting_cells(self):
        old_titles = [
            "证据编号", "是否取证", "平台", "账号类别", "发布者账号或作者", "账号 ID / UID",
            "发布时间（原页面显示）", "发布地点 / IP 属地（原页面显示）", "视频标题", "视频文案", "视频链接",
            "涉及内容 / 疑似侵权点（填写人描述）", "取证日期及时间（填写）", "取证地点（填写）", "取证文件名（填写）",
            "浏览量", "点赞数", "收藏数", "转发数", "评论数", "传播数据观察时间（填写）", "平台处理状态",
            "平台处理情况 / 投诉编号", "平台处理时间（填写）", "备注", "记录 ID", "导出开始时间（本机含时区）",
            "导出完成时间（本机含时区）", "原视频文件", "原视频 SHA-256", "处理范围（原视频时间）",
            "输出文件夹", "生成文件名", "评论区等附件文件名", "截图数量",
        ]
        with tempfile.TemporaryDirectory() as td:
            path=Path(td)/"取证台账.xlsx"
            book=Workbook()
            sheet=book.active
            sheet.title="取证台账"
            sheet.append(old_titles+["我的自定义列"])
            old_values=[f"旧值-{i}" for i in range(len(old_titles))]+["必须保留的人工内容"]
            sheet.append(old_values)
            sheet.cell(2,10,"=人工填写，不是公式")
            sheet.cell(2,len(old_titles)+1).fill=PatternFill("solid",fgColor="FFCC00")
            book.save(path)
            book.close()

            append_ledger(make_record({
                "title":"新版追加记录", "rights_guard_used":"是", "rights_guard_record_name":"录屏取证-001",
                "rights_guard_evidence_name":"某视频侵权记录", "rights_guard_evidence_id":"RG-001",
                "trusted_timestamp_used":"是", "timestamp_certificate_no":"TSA-001",
            }),td)

            book=load_workbook(path,data_only=False)
            sheet=book["取证台账"]
            headers={sheet.cell(1,column).value:column for column in range(1,sheet.max_column+1)}
            self.assertEqual(sheet.max_row,3)
            self.assertEqual([sheet.cell(1,i).value for i in range(1,len(old_titles)+1)],old_titles)
            self.assertEqual(sheet.cell(1,len(old_titles)+1).value,"我的自定义列")
            self.assertEqual(sheet.cell(2,len(old_titles)+1).value,"必须保留的人工内容")
            self.assertEqual(sheet.cell(2,10).value,"=人工填写，不是公式")
            self.assertEqual(sheet.cell(2,len(old_titles)+1).fill.fgColor.rgb,"00FFCC00")
            self.assertGreater(headers["是否使用权利卫士取证"],len(old_titles)+1)
            self.assertIsNone(sheet.cell(2,headers["是否使用权利卫士取证"]).value)
            self.assertEqual(sheet.cell(3,headers["权利卫士内记录名称"]).value,"录屏取证-001")
            self.assertEqual(sheet.cell(3,headers["权利卫士证据名称"]).value,"某视频侵权记录")
            self.assertEqual(sheet.cell(3,headers["时间戳证书编号 / 序列号"]).value,"TSA-001")
            notes=[cell[0].value for cell in book["字段说明"] if cell[0].value]
            self.assertTrue(any("旧版缺少的字段自动加到最右侧" in note for note in notes))
            book.close()


if __name__=="__main__": unittest.main()
